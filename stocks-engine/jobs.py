"""Explicit ingestion jobs. Dashboard reads never invoke these functions."""
from __future__ import annotations

import argparse
import csv
import json
import re
import urllib.error
import urllib.request
from datetime import date, datetime, timedelta
from pathlib import Path

import server
from config import settings
from providers import ArjumClient, ProviderError, YahooClient

ROOT = Path(__file__).resolve().parent
UNIVERSE_PATH = ROOT / "data" / "idx_universe.csv"

HEADER_ALIASES = {
    "symbol": {"symbol", "kode saham", "kode", "stock code", "ticker"},
    "name": {"name", "nama perusahaan", "nama emiten", "company name", "listed company"},
    "sector": {"sector", "sektor"},
    "subsector": {"subsector", "sub sektor", "subsektor"},
    "status": {"status", "listing status", "status pencatatan"},
    "instrument_type": {"instrument type", "jenis instrumen", "jenis efek", "type"},
    "board": {"board", "papan", "listing board", "papan pencatatan"},
    "listing_date": {"listing date", "tanggal pencatatan", "tanggal listing"},
    "special_notation": {"special notation", "notasi khusus"},
}
EXCLUDED_INSTRUMENTS = ("ETF", "DIRE", "REIT", "WARRANT", "WARAN", "RIGHT", "HMETD", "BOND", "SUKUK", "OBLIGASI")
INACTIVE_MARKERS = ("SUSPEND", "DELIST", "INACTIVE", "NONACTIVE", "TERHAPUS")


def normalized_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().casefold())


def canonical_columns(headers: list[object]) -> dict[str, str]:
    found = {normalized_header(header): str(header) for header in headers if header is not None}
    columns = {}
    for field, aliases in HEADER_ALIASES.items():
        match = next((found[alias] for alias in aliases if alias in found), None)
        if match:
            columns[field] = match
    missing = {"symbol", "name"} - columns.keys()
    if missing:
        raise ValueError(f"IDX export missing required columns: {sorted(missing)}. Found: {sorted(found)}")
    return columns


def read_idx_export(path: Path) -> tuple[list[dict], dict[str, str]]:
    if path.suffix.lower() in {".csv", ".txt"}:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            rows = list(reader)
            return rows, canonical_columns(reader.fieldnames or [])
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("XLSX import requires openpyxl; export IDX as CSV or install openpyxl") from exc
        book = load_workbook(path, read_only=True, data_only=True)
        sheet = book.active
        values = sheet.iter_rows(values_only=True)
        headers = next(values, ())
        columns = canonical_columns(list(headers))
        rows = [dict(zip(headers, row)) for row in values if any(value is not None for value in row)]
        return rows, columns
    raise ValueError("IDX export must be CSV or XLSX")


def normalize_idx_universe(path: Path) -> tuple[list[dict], list[dict]]:
    """Convert the official IDX export into the engine's stable, auditable schema."""
    raw_rows, columns = read_idx_export(path)
    included, excluded, seen = [], [], set()
    for raw in raw_rows:
        value = lambda field: str(raw.get(columns[field]) or "").strip() if field in columns else ""
        symbol = value("symbol").upper()
        name = value("name")
        instrument_type = value("instrument_type").upper()
        status = value("status").upper() or "ACTIVE"
        reason = None
        if not re.fullmatch(r"[A-Z]{4}", symbol):
            reason = "invalid_or_non_equity_ticker"
        elif symbol in seen:
            reason = "duplicate_symbol"
        elif any(marker in instrument_type for marker in EXCLUDED_INSTRUMENTS):
            reason = "non_common_stock"
        elif any(marker in status for marker in INACTIVE_MARKERS):
            reason = "inactive_or_delisted"
        if reason:
            excluded.append({"symbol": symbol, "name": name, "reason": reason})
            continue
        seen.add(symbol)
        included.append({
            "symbol": symbol, "name": name or symbol, "sector": value("sector"),
            "subsector": value("subsector"), "status": "ACTIVE",
            "instrument_type": instrument_type or "COMMON_STOCK_UNCLASSIFIED",
            "board": value("board"), "listing_date": value("listing_date"),
            "special_notation": value("special_notation"),
            "source_as_of": date.today().isoformat(),
        })
    if not included:
        raise ValueError("IDX export produced zero eligible common stocks")
    return included, excluded


def import_idx_universe(path: Path, allow_small: bool = False) -> None:
    rows, excluded = normalize_idx_universe(path)
    if len(rows) < settings.min_universe_size and not allow_small:
        raise ValueError(f"IDX universe has {len(rows)} eligible stocks; expected at least {settings.min_universe_size}. Use --allow-small only for fixtures.")
    with UNIVERSE_PATH.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader(); writer.writerows(rows)
    audit_path = settings.data_dir / "idx_universe_import_audit.json"
    audit_path.write_text(json.dumps({"source_file": path.name, "imported_at": date.today().isoformat(), "eligible_count": len(rows), "excluded_count": len(excluded), "excluded": excluded}, ensure_ascii=False, indent=2), encoding="utf-8")
    server.init_db(); db = server.connect()
    try:
        import engine
        engine.sync_universe(db, rows)
    finally:
        db.close()
    print(f"Imported {len(rows)} eligible IDX common stocks; excluded {len(excluded)} -> {UNIVERSE_PATH}")


def import_snapshot(path: Path) -> None:
    payload = json.loads(path.read_text(encoding="utf-8"))
    required = {"evaluation_date", "generated_at", "status", "source_mode", "rescan_on_read", "universe"}
    missing = required - payload.keys()
    if missing:
        raise ValueError(f"snapshot missing keys: {sorted(missing)}")
    if payload["rescan_on_read"] is not False:
        raise ValueError("production reports must set rescan_on_read=false")
    db = server.connect()
    db.execute("INSERT OR REPLACE INTO reports VALUES(?,?,?,?,?,?)", (
        payload["evaluation_date"], payload["status"], json.dumps(payload),
        payload["generated_at"], payload["source_mode"], 0,
    ))
    db.commit(); db.close()
    print(f"Imported {payload['evaluation_date']} from {path}")


def read_path(data: dict, *path: str):
    value = data
    for key in path:
        if not isinstance(value, dict): return None
        value = value.get(key)
    return value if isinstance(value, (int, float)) else None


def numeric_value(value):
    if isinstance(value, bool): return None
    if isinstance(value, (int, float)): return float(value)
    if isinstance(value, dict):
        for key in ("total", "value", "nilai"):
            if isinstance(value.get(key), (int, float)): return float(value[key])
    return None


def find_numeric_key(data: object, names: tuple[str, ...]):
    """Return the first exact financial-schema key in priority order."""
    for name in names:
        found = []
        def walk(value):
            if not isinstance(value, dict): return
            for key, child in value.items():
                if key == name:
                    number = numeric_value(child)
                    if number is not None: found.append(number)
                walk(child)
        walk(data)
        if found: return found[0]
    return None


def normalize_financial_statement(payload: dict) -> list[dict]:
    """Normalize Arjum financial statements without guessing missing ratios."""
    symbol, report_type = str(payload.get("stock_code") or "").upper(), str(payload.get("report_type") or "")
    if not symbol or not isinstance(payload.get("items"), list):
        raise ValueError("financial payload requires stock_code and items[]")
    normalized = []
    for item in payload["items"]:
        if not isinstance(item, dict) or not isinstance(item.get("data"), dict): continue
        data = item["data"]
        metrics = {"report_type": report_type, "period": payload.get("period"), "label": item.get("label"),
          "revenue": find_numeric_key(data, ("pendapatan", "penjualan_dan_pendapatan_usaha", "pendapatan_dan_penjualan", "pendapatan_dan_beban_operasional")),
          "operating_income": find_numeric_key(data, ("laba_operasional", "laba_rugi_usaha")),
          "net_income": find_numeric_key(data, ("laba_rugi_yang_dapat_diatribusikan_ke_entitas_induk", "laba_rugi")),
          "eps": find_numeric_key(data, ("laba_per_saham_dasar_diatribusikan_kepada_pemilik_entitas_induk", "laba_rugi_per_saham_dasar_dari_operasi_yang_dilanjutkan")),
          "total_assets": find_numeric_key(data, ("aset",)),
          "total_liabilities": find_numeric_key(data, ("liabilitas",)),
          "total_equity": find_numeric_key(data, ("ekuitas",)),
          "operating_cash_flow": find_numeric_key(data, ("arus_kas_bersih_yang_diperoleh_dari_digunakan_untuk_aktivitas_operasi",)),
          "dividend_paid": find_numeric_key(data, ("pembayaran_dividen_dari_aktivitas_pendanaan", "pembayaran_dividen", "dividen_yang_dibayarkan", "pembayaran_dividen_kas")),
          "raw": data}
        normalized.append({"symbol": symbol, "period_end": f"{item.get('year', 'UNKNOWN')}-Q{item.get('quarter', 'UNKNOWN')}", "published_at": item.get("fetched_at"), "metrics": metrics})
    if not normalized: raise ValueError("financial payload contains no usable statement items")
    return normalized


def import_fundamental(path: Path) -> None:
    payload = json.loads(path.read_text(encoding="utf-8")); rows = normalize_financial_statement(payload)
    server.init_db(); db = server.connect()
    try:
        for row in rows:
            save_financial_row(db, row, payload.get("report_type"), payload.get("period", "quarterly"))
        db.commit()
    finally: db.close()
    print(f"Imported {len(rows)} {payload.get('report_type')} periods for {rows[0]['symbol']}")


def save_financial_row(db, row: dict, report_type: str, period: str) -> None:
    db.execute("""INSERT INTO financial_statements(symbol,report_type,period,period_end,published_at,fetched_at,metrics_json,source,data_status)
      VALUES(?,?,?,?,?,?,?,?,?) ON CONFLICT(symbol,report_type,period,period_end) DO UPDATE SET
      published_at=excluded.published_at,fetched_at=excluded.fetched_at,metrics_json=excluded.metrics_json,
      source=excluded.source,data_status=excluded.data_status""",
      (row["symbol"], str(report_type).upper(), period, row["period_end"], row["published_at"], datetime.now().astimezone().isoformat(timespec="seconds"),
       json.dumps(row["metrics"], ensure_ascii=False), "arjum_financial_statement", "CACHED"))


def collect_symbol(symbol: str, include_arjum: bool) -> None:
    output = {"symbol": symbol.upper(), "providers": {}, "errors": []}
    if settings.yahoo_enabled:
        try: output["providers"]["yahoo_daily"] = YahooClient().chart(symbol, "1d", "3mo")
        except ProviderError as exc: output["errors"].append({"provider": "yahoo", "error": str(exc)})
    if include_arjum:
        if not settings.arjum_api_key:
            raise RuntimeError("ARJUM_API_KEY is required for --arjum")
        client = ArjumClient(settings.arjum_api_key)
        for name, fn in (("analysis",client.analysis),("history",client.history),("broker_summary",client.broker_summary),("broker_accumulation",client.broker_accumulation),("seasonal",client.seasonal)):
            try: output["providers"][f"arjum_{name}"] = fn(symbol)
            except ProviderError as exc: output["errors"].append({"provider": f"arjum_{name}", "error": str(exc)})
    target = settings.data_dir / "raw" / str(date.today())
    target.mkdir(parents=True, exist_ok=True)
    path = target / f"{symbol.upper()}.json"
    path.write_text(json.dumps(output, ensure_ascii=False), encoding="utf-8")
    print(f"Collected {symbol.upper()} -> {path} ({len(output['errors'])} errors)")


def collect_arjum_screener(enrich: bool = False) -> None:
    """Explicit Arjum pipeline stage; never invoked by dashboard reads."""
    import engine
    db = server.connect()
    try:
        engine.ensure_runtime(db)
        screener = engine.collect_arjum_screener(db)
        result = {"screener": screener}
        if enrich and screener["status"] in {"CACHED", "ERROR"}:
            symbols = sorted(engine.arjum_screened_symbols(db))
            result["enrichment"] = engine.collect_arjum_candidates(db, [(symbol, 100) for symbol in symbols])
        print(json.dumps(result, ensure_ascii=False, indent=2))
    finally:
        db.close()


def collect_arjum_shortlist(limit: int = 20) -> None:
    """Complete cached Arjum profiles for the highest-scored actionable plans."""
    import engine
    db = server.connect()
    try:
        engine.ensure_runtime(db)
        rows = db.execute("""SELECT symbol,MAX(COALESCE(evaluation_score,0)) score FROM instruments
          WHERE evaluation_status='ACTIONABLE' GROUP BY symbol ORDER BY score DESC,symbol LIMIT ?""", (limit,)).fetchall()
        result = engine.collect_arjum_candidates(db, [(row["symbol"], row["score"]) for row in rows])
        print(json.dumps({"selected": len(rows), "results": result}, ensure_ascii=False, indent=2))
    finally:
        db.close()


def collect_arjum_universe(limit: int | None = None) -> None:
    """Warm daily Arjum analysis and broker-summary cache for active IDX equities.

    This is an explicit maintenance job. It is resumable because complete,
    same-day profiles consume zero requests on subsequent runs.
    """
    import engine
    db = server.connect()
    try:
        engine.ensure_runtime(db)
        sql = "SELECT symbol FROM instruments WHERE status='ACTIVE' ORDER BY symbol"
        rows = db.execute(sql).fetchall()
        symbols = [row["symbol"] for row in rows]
        if limit is not None:
            symbols = symbols[:max(0, limit)]
        summary = {"selected": len(symbols), "cached": 0, "errors": 0, "quota_skips": 0, "requests": 0}
        print(json.dumps({"event": "arjum_universe_started", "selected": len(symbols)}, ensure_ascii=False), flush=True)
        for index, symbol in enumerate(symbols, 1):
            try:
                result = engine.collect_arjum_snapshot(db, symbol, engine.ARJUM_DAILY_PROFILE_ENDPOINTS, refresh_daily=True)
            except Exception as exc:
                result = {"symbol": symbol, "status": "ERROR", "requests": 0, "error": str(exc)}
            summary["requests"] += int(result.get("requests") or 0)
            if result.get("status") == "CACHED": summary["cached"] += 1
            elif result.get("status") == "SKIPPED_QUOTA": summary["quota_skips"] += 1
            else: summary["errors"] += 1
            if index % 25 == 0 or index == len(symbols):
                print(json.dumps({"event": "arjum_universe_progress", "completed": index, **summary}, ensure_ascii=False), flush=True)
            if result.get("status") == "SKIPPED_QUOTA":
                break
        print(json.dumps({"event": "arjum_universe_completed", **summary}, ensure_ascii=False), flush=True)
    finally:
        db.close()


def collect_arjum_fundamentals(limit: int | None = None, refresh: bool = False) -> None:
    """Cache quarterly income, balance-sheet, and cash-flow reports for IDX equities."""
    import engine
    report_types = ("INCOME_STATEMENT", "BALANCE_SHEET", "CASH_FLOW_REPORT")
    db = server.connect()
    try:
        engine.ensure_runtime(db)
        symbols = [row[0] for row in db.execute("SELECT symbol FROM instruments WHERE status='ACTIVE' ORDER BY symbol")]
        if limit is not None: symbols = symbols[:max(0, limit)]
        summary = {"selected":len(symbols),"reports_cached":0,"reports_skipped":0,"errors":0,"requests":0,"quota_skips":0}
        client = ArjumClient(settings.arjum_api_key or "")
        print(json.dumps({"event":"arjum_fundamentals_started","selected":len(symbols),"report_types":len(report_types)},ensure_ascii=False),flush=True)
        stop = False
        for index, symbol in enumerate(symbols, 1):
            for report_type in report_types:
                cached = db.execute("""SELECT fetched_at FROM financial_statements WHERE symbol=? AND report_type=? AND period='quarterly'
                  ORDER BY fetched_at DESC LIMIT 1""",(symbol,report_type)).fetchone()
                if cached and not refresh:
                    try: fresh = datetime.fromisoformat(cached[0]).date() == date.today()
                    except (TypeError, ValueError): fresh = False
                    if fresh:
                        summary["reports_skipped"] += 1; continue
                used, ceiling = engine.arjum_usage(db)
                if used >= ceiling:
                    summary["quota_skips"] += 1; stop = True; break
                today = engine.now_wib().date().isoformat()
                db.execute("UPDATE provider_usage SET requests_used=requests_used+1 WHERE provider='arjum' AND usage_date=?",(today,));db.commit()
                summary["requests"] += 1
                try:
                    payload = client.financial_statement(symbol, report_type, "quarterly", 12)
                    rows = normalize_financial_statement(payload)
                    for row in rows: save_financial_row(db,row,report_type,"quarterly")
                    db.commit(); summary["reports_cached"] += 1
                except (ProviderError, ValueError) as exc:
                    summary["errors"] += 1
                    engine.log_event(db,"ERROR","arjum_financial","report_failed",{"symbol":symbol,"report_type":report_type,"error":str(exc)})
                    db.commit()
            if index % 10 == 0 or index == len(symbols) or stop:
                print(json.dumps({"event":"arjum_fundamentals_progress","completed":index,**summary},ensure_ascii=False),flush=True)
            if stop: break
        print(json.dumps({"event":"arjum_fundamentals_completed",**summary},ensure_ascii=False),flush=True)
    finally:
        db.close()


def refresh_cached_scores() -> None:
    """Replace placeholder scores using the latest completed cached candle only."""
    import engine
    db=server.connect(); updated=missing=0
    try:
        engine.ensure_runtime(db)
        symbols=[row[0] for row in db.execute("SELECT symbol FROM instruments WHERE status='ACTIVE' ORDER BY symbol")]
        for symbol in symbols:
            features=None
            for timeframe in ("1d","5m","15m","1h","1m"):
                features=engine.feature_set(engine.candle_rows(db,symbol,timeframe))
                if features: break
            if not features:
                missing += 1; continue
            score=engine.technical_score(features)
            db.execute("UPDATE instruments SET evaluation_score=?,evaluation_status=CASE WHEN evaluation_status='ACTIONABLE' THEN evaluation_status ELSE 'WATCH' END WHERE symbol=?",(score,symbol))
            updated += 1
        engine.log_event(db,"INFO","score_backfill","cached_scores_refreshed",{"updated":updated,"missing":missing,"external_requests":0})
        db.commit()
    finally: db.close()
    print(json.dumps({"status":"COMPLETED","updated":updated,"missing":missing,"external_requests":0},ensure_ascii=False))


def validate_stock_pages(base_url: str = "http://127.0.0.1:4175") -> None:
    """Validate every active stock API payload against the stock-page contract."""
    db = server.connect()
    try:
        symbols = [row[0] for row in db.execute("SELECT symbol FROM instruments WHERE status='ACTIVE' ORDER BY symbol")]
    finally:
        db.close()
    failures, status_counts = [], {}
    array_fields = ("agent_views", "positions", "decision_history", "report_history")
    for index, symbol in enumerate(symbols, 1):
        issues = []
        try:
            request = urllib.request.Request(f"{base_url.rstrip('/')}/api/stocks/{symbol}", headers={"Accept": "application/json"})
            with urllib.request.urlopen(request, timeout=15) as response:
                if response.status != 200: issues.append(f"http_{response.status}")
                payload = json.load(response)
            if payload.get("symbol") != symbol: issues.append("symbol_mismatch")
            for field in array_fields:
                if not isinstance(payload.get(field), list): issues.append(f"{field}_not_array")
            chart = payload.get("chart")
            if not isinstance(chart, dict):
                issues.append("chart_not_object")
            else:
                available = chart.get("available_timeframes")
                series = chart.get("series")
                if not isinstance(available, list): issues.append("available_timeframes_not_array")
                if not isinstance(series, dict): issues.append("chart_series_not_object")
                if isinstance(available, list) and isinstance(series, dict):
                    for timeframe in available:
                        if not isinstance(series.get(timeframe), list): issues.append(f"series_{timeframe}_not_array")
            for view in payload.get("agent_views") or []:
                plan = view.get("plan") or {}
                if not isinstance(plan.get("targets"), list): issues.append(f"{view.get('agent_id')}_targets_not_array")
            flow = payload.get("flow") or {}
            if flow.get("status") == "AVAILABLE":
                if not isinstance(flow.get("heavy_buyers"), list): issues.append("heavy_buyers_not_array")
                if not isinstance(flow.get("heavy_sellers"), list): issues.append("heavy_sellers_not_array")
            arjum = payload.get("arjum") or {}
            status_counts[arjum.get("status", "MISSING")] = status_counts.get(arjum.get("status", "MISSING"), 0) + 1
            if arjum.get("status") == "CACHED":
                if not isinstance(arjum.get("pivot_support"), list): issues.append("pivot_support_not_array")
                if not isinstance(arjum.get("pivot_resistance"), list): issues.append("pivot_resistance_not_array")
                if not isinstance(arjum.get("payload"), dict): issues.append("arjum_payload_not_object")
        except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as exc:
            issues.append(f"request_error:{exc}")
        except Exception as exc:
            issues.append(f"validation_error:{type(exc).__name__}:{exc}")
        if issues: failures.append({"symbol": symbol, "issues": sorted(set(issues))})
        if index % 100 == 0 or index == len(symbols):
            print(json.dumps({"event":"stock_page_validation_progress","completed":index,"total":len(symbols),"failures":len(failures)}, ensure_ascii=False), flush=True)
    report = {"validated_at": date.today().isoformat(), "base_url": base_url, "total": len(symbols), "passed": len(symbols)-len(failures), "failed": len(failures), "arjum_status": status_counts, "failures": failures}
    target = settings.data_dir / "stock_page_validation.json"
    target.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({**report, "report_path": str(target)}, ensure_ascii=False, indent=2))


def reset_paper_agents() -> None:
    """Reset only simulated portfolio state; market caches and universe remain intact."""
    server.init_db(); db = server.connect()
    timestamp = f"{date.today().isoformat()}T00:00:00+07:00"
    try:
        db.execute("BEGIN IMMEDIATE")
        # Dependent records first; retained candles and Arjum snapshots allow
        # the restarted engine to evaluate immediately without a fresh scan.
        for table in ("paper_fills", "paper_orders", "positions", "trade_journal", "agent_proposals", "decisions", "equity_history"):
            db.execute(f"DELETE FROM {table}")
        db.execute("UPDATE agents SET starting_equity=100000000,equity=100000000,win_rate=0,open_risk_pct=0")
        agent_ids = [row[0] for row in db.execute("SELECT id FROM agents")]
        for agent_id in agent_ids:
            db.execute("""INSERT INTO agent_ledgers(agent_id,cash,realized_pnl,fees_paid,updated_at) VALUES(?,?,?,?,?)
              ON CONFLICT(agent_id) DO UPDATE SET cash=excluded.cash,realized_pnl=excluded.realized_pnl,fees_paid=excluded.fees_paid,updated_at=excluded.updated_at""",
              (agent_id, 100_000_000, 0, 0, timestamp))
            db.execute("INSERT INTO equity_history(agent_id,equity_date,equity,cash,drawdown_pct) VALUES(?,?,?,?,?)",
              (agent_id, date.today().isoformat(), 100_000_000, 100_000_000, 0))
        db.execute("INSERT INTO engine_events(created_at,level,component,event,details_json) VALUES(?,?,?,?,?)",
          (timestamp, "INFO", "paper_ledger", "agents_reset", json.dumps({"agents": len(agent_ids), "equity_each": 100_000_000})))
        db.commit()
    except Exception:
        db.rollback(); raise
    finally:
        db.close()
    print(json.dumps({"status": "RESET", "agents": len(agent_ids), "equity_each": 100_000_000, "positions": 0, "pending_orders": 0}, ensure_ascii=False))


def main() -> None:
    parser = argparse.ArgumentParser(description="NusaQuant explicit data jobs")
    sub = parser.add_subparsers(dest="command", required=True)
    imp = sub.add_parser("import-snapshot"); imp.add_argument("path", type=Path)
    fundamental = sub.add_parser("import-fundamental", help="import an Arjum financial-statement JSON response"); fundamental.add_argument("path", type=Path)
    universe = sub.add_parser("import-idx-universe", help="normalize official IDX CSV/XLSX Stock List export")
    universe.add_argument("path", type=Path)
    universe.add_argument("--allow-small", action="store_true", help="only for fixtures; production requires the configured minimum universe")
    collect = sub.add_parser("collect-symbol"); collect.add_argument("symbol"); collect.add_argument("--arjum", action="store_true")
    arjum = sub.add_parser("collect-arjum-screener", help="cache Arjum screener; --enrich fetches analysis + broker summary for its rows")
    arjum.add_argument("--enrich", action="store_true")
    shortlist = sub.add_parser("collect-arjum-shortlist", help="complete Arjum cache for actionable plans")
    shortlist.add_argument("--limit", type=int, default=20)
    full_arjum = sub.add_parser("collect-arjum-universe", help="refresh daily Arjum analysis and broker summary for all active IDX stocks")
    full_arjum.add_argument("--limit", type=int, default=None, help="optional deterministic prefix for a test run")
    financials = sub.add_parser("collect-arjum-fundamentals", help="cache quarterly income, balance-sheet, and cash-flow reports")
    financials.add_argument("--limit", type=int, default=None, help="optional deterministic prefix for a test run")
    financials.add_argument("--refresh", action="store_true", help="replace reports already fetched today")
    sub.add_parser("refresh-cached-scores", help="replace placeholder scores from completed cached candles; no provider requests")
    validate = sub.add_parser("validate-stock-pages", help="request and validate every active stock detail API")
    validate.add_argument("--base-url", default="http://127.0.0.1:4175")
    sub.add_parser("reset-paper-agents", help="clear simulated orders, positions, journals, and reset all agents to Rp100m")
    args = parser.parse_args()
    server.init_db()
    if args.command == "import-snapshot": import_snapshot(args.path)
    elif args.command == "import-fundamental": import_fundamental(args.path)
    elif args.command == "import-idx-universe": import_idx_universe(args.path, args.allow_small)
    elif args.command == "collect-symbol": collect_symbol(args.symbol, args.arjum)
    elif args.command == "collect-arjum-screener": collect_arjum_screener(args.enrich)
    elif args.command == "collect-arjum-shortlist": collect_arjum_shortlist(args.limit)
    elif args.command == "collect-arjum-universe": collect_arjum_universe(args.limit)
    elif args.command == "collect-arjum-fundamentals": collect_arjum_fundamentals(args.limit, args.refresh)
    elif args.command == "refresh-cached-scores": refresh_cached_scores()
    elif args.command == "validate-stock-pages": validate_stock_pages(args.base_url)
    elif args.command == "reset-paper-agents": reset_paper_agents()


if __name__ == "__main__": main()
